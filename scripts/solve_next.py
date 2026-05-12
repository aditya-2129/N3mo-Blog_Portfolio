#!/usr/bin/env python3
"""
Find the next To Do Easy problem in the Excel sheet and solve it.

Usage:
    python scripts/solve_next.py              # next Easy To Do
    python scripts/solve_next.py --difficulty Medium
    python scripts/solve_next.py --difficulty Hard
    python scripts/solve_next.py --count 3    # solve 3 Easy problems in a row
"""

import sys
import re
import os
import time
import argparse
import subprocess
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).parent
ROOT_DIR = SCRIPT_DIR.parent
XLSX_PATH = SCRIPT_DIR / "LeetCode Questions - Last Updated 2026-05-08 05_09 CEST.xlsx"

load_dotenv(ROOT_DIR / ".env")


def find_todo_problems(difficulty: str, count: int) -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Find header row
    header_row = None
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == "ID":
            header_row = i
            break
    if header_row is None:
        raise SystemExit("ERROR: Could not find header row in Excel sheet.")

    # Find Status column
    status_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == "Status":
            status_col = c
            break
    if status_col is None:
        raise SystemExit("ERROR: Status column not found. Run the main script first to add it.")

    found = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=1).value
        if raw_id is None:
            continue
        try:
            problem_id = int(raw_id)
        except (ValueError, TypeError):
            continue

        diff = str(ws.cell(row=row_idx, column=7).value or "").strip()
        status = str(ws.cell(row=row_idx, column=status_col).value or "").strip()

        if diff.lower() == difficulty.lower() and status != "Done":
            # Extract problem name from HYPERLINK formula
            name_raw = ws.cell(row=row_idx, column=2).value or ""
            name_raw = str(name_raw)
            m = re.search(r'"([^"]+)"\s*\)\s*$', name_raw)
            name = m.group(1) if m else name_raw

            found.append({"id": problem_id, "name": name, "difficulty": diff})
            if len(found) >= count:
                break

    return found


def main():
    parser = argparse.ArgumentParser(
        description="Solve the next To Do problem(s) from the Excel sheet."
    )
    parser.add_argument(
        "--difficulty", "-d",
        default="Easy",
        choices=["Easy", "Medium", "Hard"],
        help="Difficulty to filter (default: Easy)",
    )
    parser.add_argument(
        "--count", "-n",
        type=int,
        default=1,
        help="How many problems to solve in sequence (default: 1)",
    )
    args = parser.parse_args()

    if not os.environ.get("GEMINI_API_KEY"):
        raise SystemExit(
            "ERROR: GEMINI_API_KEY is not set.\n"
            "  Add it to your .env file or run: $env:GEMINI_API_KEY = 'AIza...'"
        )

    print(f"\nSearching for {args.count} To Do '{args.difficulty}' problem(s) in Excel...")
    problems = find_todo_problems(args.difficulty, args.count)

    if not problems:
        print(f"No To Do '{args.difficulty}' problems found — you've solved them all!")
        return

    print(f"Found {len(problems)} problem(s) to solve:")
    for p in problems:
        print(f"  #{p['id']} — {p['name']}")
    print()

    solve_script = SCRIPT_DIR / "solve.py"
    success, failed = [], []

    for i, problem in enumerate(problems, 1):
        print(f"[{i}/{len(problems)}] Solving #{problem['id']}: {problem['name']} ({problem['difficulty']})")
        print("-" * 60)

        result = subprocess.run(
            [sys.executable, str(solve_script), str(problem["id"])],
            cwd=str(ROOT_DIR),
        )

        if result.returncode == 0:
            success.append(problem)
        else:
            failed.append(problem)
            print(f"  FAILED: #{problem['id']} — skipping to next.\n")

        # Brief pause between API calls to avoid rate limiting
        if i < len(problems):
            print("  Waiting 10s before next problem...")
            time.sleep(10)

    print("\n" + "=" * 60)
    print(f"Done! {len(success)}/{len(problems)} solved successfully.")
    if success:
        print("  Solved:")
        for p in success:
            print(f"    #{p['id']} — {p['name']}")
    if failed:
        print("  Failed (retry manually with: python scripts/solve.py <id>):")
        for p in failed:
            print(f"    #{p['id']} — {p['name']}")


if __name__ == "__main__":
    main()
