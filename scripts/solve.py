#!/usr/bin/env python3
"""
Generate a LeetCode blog post from the Excel sheet and update its status.

Usage:
    python scripts/solve.py <problem_id>
    python scripts/solve.py <problem_id> --provider nvidia

Example:
    python scripts/solve.py 29
    python scripts/solve.py 29 --provider nvidia --test

Requirements:
    pip install google-generativeai openpyxl openai

Setup (one-time):
    Gemini : $env:GEMINI_API_KEY = "AIza..."
    NVIDIA : $env:NVIDIA_API_KEY = "nvapi-..."
"""

import sys
import re
import os
import json
import argparse
from pathlib import Path
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from google import genai
from google.genai import types as genai_types
from dotenv import load_dotenv

# ── Paths (resolved relative to this script's location) ────────────────────
SCRIPT_DIR = Path(__file__).parent
ROOT_DIR = SCRIPT_DIR.parent

# Load .env from project root
load_dotenv(ROOT_DIR / ".env")
XLSX_PATH = SCRIPT_DIR / "LeetCode Questions - Last Updated 2026-05-08 05_09 CEST.xlsx"
BLOG_DIR = ROOT_DIR / "src" / "content" / "blog"
TEST_DIR = ROOT_DIR / "scripts" / "model_comparison"

# ── Gemini model aliases ──────────────────────────────────────────────────────
GEMINI_MODELS = {
    "flash":      "gemini-3-flash-preview",
    "2.5-flash":  "gemini-2.5-flash",
    "lite":       "gemini-3.1-flash-lite",
    "2.5-lite":   "gemini-2.5-flash-lite",
}
DEFAULT_GEMINI_MODEL = "flash"

# ── NVIDIA NIM config ────────────────────────────────────────────────────────
NVIDIA_BASE_URL = "https://integrate.api.nvidia.com/v1"
NVIDIA_MODEL = "z-ai/glm4.7"
NVIDIA_MAX_TOKENS = 32768

# ── Providers ─────────────────────────────────────────────────────────────
PROVIDERS = ["gemini", "nvidia"]
DEFAULT_PROVIDER = "nvidia"

# ── 5 languages (covers 90%+ of LeetCode users) ────────────────────────────
LANGUAGES = ["Python3", "Java", "C++", "JavaScript", "TypeScript"]
LANG_ATTR = {
    "Python3": "python", "Java": "java", "C++": "cpp",
    "JavaScript": "javascript", "TypeScript": "typescript",
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def slugify(s: str) -> str:
    s = str(s).lower().strip()
    s = re.sub(r"'", "", s)       # drop apostrophes: Pascal's -> pascals
    s = re.sub(r"[^a-z0-9\s-]", "-", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def extract_hyperlink_text(raw) -> str:
    """Pull display name from =HYPERLINK("url","Name") formula."""
    if raw is None:
        return ""
    s = str(raw)
    if s.startswith("=HYPERLINK"):
        m = re.search(r'"([^"]+)"\s*\)\s*$', s)
        return m.group(1) if m else s
    return s


def esc_backticks(s: str) -> str:
    return str(s).replace("`", "\\`")


def esc_mdx(s: str) -> str:
    """Escape < and > in plain text so MDX doesn't treat them as JSX tags."""
    return str(s).replace("<", "&lt;").replace(">", "&gt;")


# ── Excel ─────────────────────────────────────────────────────────────────────

def load_problem(problem_id: int) -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Find the header row (the one that starts with "ID")
    header_row = None
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == "ID":
            header_row = i
            break
    if header_row is None:
        raise SystemExit("ERROR: Could not find header row in Excel sheet.")

    # Find Status column index
    status_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == "Status":
            status_col = c
            break

    # Scan for the problem row
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=1).value
        if raw_id is None:
            continue
        try:
            if int(raw_id) != problem_id:
                continue
        except (ValueError, TypeError):
            continue

        name = extract_hyperlink_text(ws.cell(row=row_idx, column=2).value)
        topics_raw = ws.cell(row=row_idx, column=6).value or ""
        topics = [t.strip() for t in str(topics_raw).split(",") if t.strip()]
        difficulty = str(ws.cell(row=row_idx, column=7).value or "")
        accepted = ws.cell(row=row_idx, column=8).value
        submissions = ws.cell(row=row_idx, column=9).value
        accept_rate = ws.cell(row=row_idx, column=10).value
        is_free = ws.cell(row=row_idx, column=11).value == "Yes"

        if accept_rate:
            try:
                accept_pct = f"{float(accept_rate) * 100:.2f}%"
            except (ValueError, TypeError):
                accept_pct = str(accept_rate)
        else:
            accept_pct = "N/A"

        return {
            "id": problem_id,
            "name": name,
            "difficulty": difficulty,
            "topics": topics,
            "accept_pct": accept_pct,
            "is_free": is_free,
            "row_idx": row_idx,
            "status_col": status_col,
            "wb": wb,
            "ws": ws,
        }

    raise SystemExit(f"ERROR: Problem ID {problem_id} not found in Excel sheet.")


def mark_done(problem: dict):
    ws = problem["ws"]
    wb = problem["wb"]
    status_col = problem["status_col"]
    if not status_col:
        print("WARNING: Status column not found - Excel not updated.")
        return

    cell = ws.cell(row=problem["row_idx"], column=status_col)
    cell.value = "Done"
    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    cell.font = Font(color="276221", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(XLSX_PATH)
    print(f"  Excel updated -> Problem {problem['id']} marked Done")


# ── Claude API ────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """\
You are an expert competitive programmer and technical writer.
Always respond with a single valid JSON object - no markdown fences, no prose outside the JSON.
Backtick characters (`) inside code strings must be replaced with a single quote (') \
so the output is safe to embed in MDX template literals."""


def build_user_prompt(problem: dict) -> str:
    topics = ", ".join(problem["topics"]) if problem["topics"] else "N/A"
    langs = ", ".join(LANGUAGES)
    return f"""\
Generate a complete LeetCode editorial for:

  ID         : {problem['id']}
  Title      : {problem['name']}
  Difficulty : {problem['difficulty']}
  Topics     : {topics}

Return ONLY a JSON object with this exact shape:

{{
  "Problem": "<exact problem title>",
  "Description": "<full LeetCode problem statement - preserve math notation>",
  "Examples": [
    {{
      "Input": "<example input>",
      "Output": "<example output>",
      "Explanation": "<explanation or empty string>"
    }}
  ],
  "Constraints": ["<constraint1>", "<constraint2>"],
  "Approaches": [
    {{
      "name": "<Approach Name>",
      "First Intuition": "<1-2 sentence high-level idea>",
      "Problem-Solving Approach": ["<step 1>", "<step 2>", "..."],
      "Code Implementation": {{
        "Python3": "<working Python 3 code with type hints>",
        "Java": "<working Java code>",
        "C++": "<working C++ code>",
        "JavaScript": "<working JS code>",
        "TypeScript": "<working TypeScript code>"
      }},
      "Complexity Analysis": {{
        "Time": "<big-O with brief reason>",
        "Space": "<big-O with brief reason>",
        "Notes": "<trade-offs or caveats>"
      }}
    }}
  ]
}}

Rules:
- Include ALL meaningful approaches that exist for this problem (brute-force through every known optimal).
  Do NOT artificially cap at 2 or 3 — if 5 approaches exist (e.g. brute force, two pointers, binary search,
  DP, bit manipulation), include all 5. Think like a competitive programmer writing a full editorial.
- All 5 languages ({langs}) must be present in EVERY approach.
- Code must be complete and runnable (include class/function signatures LeetCode expects).
- Do NOT use backtick characters anywhere in code strings.
- For math exponents use Unicode superscripts, NOT caret notation.
  Examples: write 2³¹ not 2^31, write n² not n^2, write 10⁶ not 10^6, write O(n²) not O(n^2).
  Unicode superscript digits: ⁰¹²³⁴⁵⁶⁷⁸⁹. Negative sign in exponent: use ⁻ (e.g. 2⁻¹).
  Superscript letters for Big-O: ⁿ (e.g. 2ⁿ).
- Examples must match the actual LeetCode problem exactly.
- Return only the JSON object - nothing else.
"""


def call_gemini(problem: dict, model_key: str = DEFAULT_GEMINI_MODEL) -> dict:
    client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))
    model_id = GEMINI_MODELS.get(model_key, model_key)

    print(f"  Calling {model_id} for problem {problem['id']}: {problem['name']} ...")

    for attempt in range(1, 4):
        response = client.models.generate_content(
            model=model_id,
            contents=build_user_prompt(problem),
            config=genai_types.GenerateContentConfig(
                system_instruction=SYSTEM_PROMPT,
                temperature=0.2,
                max_output_tokens=65536,
            ),
        )

        finish = getattr(response.candidates[0], "finish_reason", None) if response.candidates else None
        raw = response.text.strip()

        # Strip accidental markdown fences
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            if str(finish) in ("MAX_TOKENS", "2") and attempt < 3:
                print(f"  Response truncated (attempt {attempt}), retrying with shorter prompt ...")
                continue
            m = re.search(r"\{[\s\S]*\}", raw)
            if m:
                try:
                    return json.loads(m.group(0))
                except json.JSONDecodeError:
                    pass
            raise SystemExit(
                f"ERROR: Gemini returned invalid JSON after {attempt} attempt(s).\n"
                f"finish_reason={finish}\nRaw (first 800 chars):\n{raw[:800]}"
            )


# ── MDX Renderer ──────────────────────────────────────────────────────────────

def render_codetabs(code_impl: dict) -> str:
    tabs = []
    for lang in LANGUAGES:
        code = code_impl.get(lang, "")
        if not code:
            continue
        lang_attr = LANG_ATTR.get(lang, lang.lower())
        tabs.append(
            f'{{ name: "{lang}", lang: "{lang_attr}", code: `{esc_backticks(code)}` }}'
        )
    joined = ",\n".join(tabs)
    return f"<CodeTabs client:load tabs={{[{joined}]}} />"


def render_mdx(doc: dict, problem: dict) -> str:
    now = datetime.now()
    date_str = now.strftime("%b %d, %Y").lstrip("0")  # "May 11, 2026"
    # Remove leading zero from day
    date_str = re.sub(r" 0(\d)", r" \1", date_str)

    title = problem["name"]
    difficulty = problem["difficulty"]
    topics = problem["topics"]
    accept_pct = problem["accept_pct"]
    paid_str = "No" if problem["is_free"] else "Yes"

    raw_desc = doc.get("Description", "")
    full_desc = re.sub(r"\s+", " ", raw_desc).strip()

    # summary is only used in frontmatter (SEO/meta) — keep it short
    if len(full_desc) > 160:
        cut = full_desc[:160]
        last_period = max(cut.rfind(". "), cut.rfind(".\n"))
        summary = cut[:last_period + 1] if last_period > 60 else cut[:157] + "..."
    else:
        summary = full_desc
    summary_escaped = summary.replace('"', '\\"')

    # full_desc goes into the article body so the complete question is visible
    full_desc_escaped = esc_mdx(full_desc)

    tags_yaml = "\n".join(f"  - {t}" for t in (topics or [difficulty]))

    # Table of contents
    approach_toc = "\n".join(
        f"- [{a['name']}](#{slugify(a['name'])})"
        for a in doc.get("Approaches", [])
    )

    # Examples
    examples_md = ""
    for ex in doc.get("Examples", []):
        examples_md += f"**Input**\n\n```text\n{ex.get('Input', '')}\n```\n\n"
        examples_md += f"**Output**\n\n```text\n{ex.get('Output', '')}\n```\n\n"
        expl = ex.get("Explanation", "").strip()
        if expl:
            examples_md += f"**Explanation**\n\n{expl}\n\n"

    # Constraints — inside a ```text``` code block so < is safe, no escaping needed
    constraints = doc.get("Constraints", [])
    if isinstance(constraints, list):
        constraints_block = "\n".join(f"- {c}" for c in constraints)
    else:
        constraints_block = str(constraints)

    # Approaches
    approaches_md = ""
    for approach in doc.get("Approaches", []):
        approaches_md += f"---\n\n## {approach['name']}\n\n"
        intuition = esc_mdx(approach.get("First Intuition", "").strip())
        if intuition:
            approaches_md += f"**Intuition**\n\n{intuition}\n\n"
        steps = approach.get("Problem-Solving Approach", [])
        if steps:
            approaches_md += "**Steps**\n\n"
            approaches_md += "\n".join(f"- {esc_mdx(s)}" for s in steps)
            approaches_md += "\n\n"
        code_impl = approach.get("Code Implementation", {})
        if code_impl:
            approaches_md += render_codetabs(code_impl) + "\n\n"
        comp = approach.get("Complexity Analysis", {})
        if comp:
            approaches_md += "**Complexity**\n\n"
            approaches_md += f"- Time: {esc_mdx(comp.get('Time', ''))}\n"
            approaches_md += f"- Space: {esc_mdx(comp.get('Space', ''))}\n"
            notes = esc_mdx(comp.get("Notes", "").strip())
            if notes:
                approaches_md += f"- Notes: {notes}\n"
            approaches_md += "\n"

    # JSON-LD
    ld_article = {
        "@context": "https://schema.org",
        "@type": "TechArticle",
        "headline": f"LeetCode {title} Solutions",
        "datePublished": now.isoformat(),
        "description": summary,
        "keywords": ", ".join(
            filter(None, ["LeetCode", title, difficulty, *topics, "Hash Map", "Brute Force", "Array"])
        ),
    }
    ld_faq = {
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": [
            {
                "@type": "Question",
                "name": f"{a['name']} complexity?",
                "acceptedAnswer": {
                    "@type": "Answer",
                    "text": (
                        f"Time: {(a.get('Complexity Analysis') or {}).get('Time', '')}; "
                        f"Space: {(a.get('Complexity Analysis') or {}).get('Space', '')}. "
                        f"{(a.get('Complexity Analysis') or {}).get('Notes', '')}"
                    ),
                },
            }
            for a in doc.get("Approaches", [])
        ],
    }

    # Build the JSX expression format matching existing posts:
    #   {JSON.stringify({...}).replace(/</g,'\\u003c')}
    def jsonld_expr(obj):
        inner = json.dumps(obj, separators=(",", ":"))
        return "{" + f"JSON.stringify({inner}).replace(/</g,'\\\\u003c')" + "}"

    mdx = f"""\
---
title: "{title}"
summary: "{summary_escaped}"
date: "{date_str}"
tags:
{tags_yaml}
draft: false
difficulty: "{difficulty}"
---

import CodeTabs from "@components/CodeTabs"

> **Difficulty:** {difficulty} | **Acceptance:** {accept_pct} | **Paid:** {paid_str}

> **Topics:** {", ".join(topics)}

{full_desc_escaped}

- Examples
- Constraints
{approach_toc}

## Examples

{examples_md}## Constraints

```text
{constraints_block}
```
---

{approaches_md}
<script is:inline type="application/ld+json">
{jsonld_expr(ld_article)}
</script>

<script is:inline type="application/ld+json">
{jsonld_expr(ld_faq)}
</script>
"""
    return mdx


# ── NVIDIA NIM Provider ────────────────────────────────────────────────────

def build_nvidia_prompt(problem: dict) -> str:
    topics = ", ".join(problem["topics"]) if problem["topics"] else "N/A"
    langs = ", ".join(LANGUAGES)
    lang_tabs = "\n".join(
        f'{{ name: "{lang}", lang: "{LANG_ATTR[lang]}", code: `<complete runnable {lang} code>` }},'
        for lang in LANGUAGES
    )
    # Generate a deterministic "publication date" from problem ID
    # Spread across Jan 2024 – Apr 2026 so posts look organic
    import random as _rnd
    _rng = _rnd.Random(problem["id"])
    _start = datetime(2024, 1, 1)
    _end = datetime(2026, 4, 30)
    _days_range = (_end - _start).days
    pub_date = _start + timedelta(days=_rng.randint(0, _days_range))
    today = pub_date.strftime("%b %d, %Y")      # e.g. "Mar 15, 2025"
    iso_date = pub_date.strftime("%Y-%m-%d")     # e.g. "2025-03-15"

    return f"""You are an expert competitive programmer and technical writer.
Always respond with the complete MDX file content only — no markdown fences wrapping the output, no commentary before or after, no "Here is your file" preamble.

Generate a complete Astro MDX blog post for this LeetCode problem:

  ID         : {problem['id']}
  Title      : {problem['name']}
  Difficulty : {problem['difficulty']}
  Topics     : {topics}
  Acceptance : {problem.get('acceptance', 'N/A')}

The MDX file MUST follow this exact structure, in this exact order:

PART 1 — FRONTMATTER:
---
title: "<exact problem title>"
summary: "<1-2 sentence problem description, max 160 chars>"
date: "{today}"
tags:
  - <Topic1>
  - <Topic2>
draft: false
difficulty: "{problem['difficulty']}"
---

PART 2 — IMPORT:
import CodeTabs from "@components/CodeTabs"

PART 3 — METADATA (as blockquotes):
> **Difficulty:** {problem['difficulty']} | **Acceptance:** {problem.get('acceptance', 'N/A')} | **Paid:** No
> **Topics:** {topics}

PART 4 — PROBLEM DESCRIPTION:
The full LeetCode problem statement as plain text.

PART 5 — TABLE OF CONTENTS:
- Examples
- Constraints
- [Approach 1 Name](#slug)
- [Approach 2 Name](#slug)
- ...etc

PART 6 — EXAMPLES: (## Examples, each with Input/Output/Explanation in ```text blocks)

PART 7 — CONSTRAINTS: (## Constraints in a ```text block)
---

PART 8 — APPROACHES:
For EACH approach (include ALL meaningful approaches — brute force through every known optimal):

---
## <Approach Name>

**Intuition**
<1-2 sentence idea>

**Steps**
- <step 1>
- <step 2>

<CodeTabs client:load tabs={{[{lang_tabs}]}} />

**Complexity**
- Time: <big-O>
- Space: <big-O>
- Notes: <trade-offs>

PART 9 — JSON-LD (at the very end of the file):
<script is:inline type="application/ld+json">
{{JSON.stringify({{"@context":"https://schema.org","@type":"TechArticle","headline":"LeetCode {problem['name']} Solutions","datePublished":"{iso_date}","description":"<summary>","keywords":"LeetCode, {problem['name']}, {problem['difficulty']}, {topics}"}}).replace(/</g,'\\u003c')}}
</script>

<script is:inline type="application/ld+json">
{{JSON.stringify({{"@context":"https://schema.org","@type":"FAQPage","mainEntity":[...array of approach complexity Q&As...]}}).replace(/</g,'\\u003c')}}
</script>

STRICT RULES:
1. Include ALL {len(LANGUAGES)} languages ({langs}) in EVERY approach. No exceptions.
2. Code must be complete and runnable on LeetCode — include class/function signatures.
3. Do NOT use backtick characters (`) anywhere inside code strings.
4. For math exponents use Unicode superscripts: 2³¹ not 2^31, n² not n^2. Digits: ⁰¹²³⁴⁵⁶⁷⁸⁹ⁿ⁻
5. Examples must match the actual LeetCode problem exactly.
6. Use language-specific idioms: Integer.MAX_VALUE (Java), INT_MAX (C++), float('inf') (Python), etc.
7. Handle all edge cases in code.
8. Include ALL meaningful approaches (do NOT cap at 2 or 3).
9. Return ONLY the MDX file content — nothing else."""


def sanitize_mdx(content: str) -> str:
    """Fix common LLM output issues that break MDX/Acorn parsing."""

    # 1. Fix HTML entities inside code template literals (inside CodeTabs)
    #    Models sometimes output &lt; &gt; &amp; instead of < > & in code strings
    def fix_code_entities(match):
        code = match.group(1)
        code = code.replace("&lt;", "<")
        code = code.replace("&gt;", ">")
        code = code.replace("&amp;", "&")
        return f"code: `{code}`"

    content = re.sub(r'code:\s*`((?:[^`]|\\`)*?)`', fix_code_entities, content)

    # 2. Escape bare < and > in plain text lines that would break MDX
    #    e.g.  "0 <= n <= 10⁵"  ->  MDX sees < as JSX tag opener
    lines = content.split("\n")
    result = []
    in_fence = False
    for line in lines:
        stripped = line.strip()

        # Track code fence state
        if stripped.startswith("```"):
            in_fence = not in_fence
            result.append(line)
            continue

        # Don't touch lines inside code fences, CodeTabs, scripts, or blockquotes
        if in_fence or stripped.startswith("<CodeTabs") or stripped.startswith("<script") or stripped.startswith(">"):
            result.append(line)
            continue

        # Escape bare < and > that aren't part of JSX/HTML tags
        # Match < not followed by a tag name or / (closing tag)
        line = re.sub(r'<(?![A-Za-z/!])', '&lt;', line)
        # Match > not preceded by a tag name or quote (end of JSX)
        line = re.sub(r'(?<![A-Za-z"\'/\-])>', '&gt;', line)

        result.append(line)

    return "\n".join(result)


def call_nvidia(problem: dict) -> str:
    try:
        from openai import OpenAI
    except ImportError:
        raise SystemExit(
            "ERROR: openai package not installed.\n"
            "  Run: pip install openai"
        )

    client = OpenAI(
        base_url=NVIDIA_BASE_URL,
        api_key=os.environ.get("NVIDIA_API_KEY"),
    )

    print(f"  Calling NVIDIA NIM ({NVIDIA_MODEL}) for problem {problem['id']}: {problem['name']} ...")

    prompt = build_nvidia_prompt(problem)

    for attempt in range(1, 4):
        try:
            completion = client.chat.completions.create(
                model=NVIDIA_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=1,
                top_p=1,
                max_tokens=NVIDIA_MAX_TOKENS,
                seed=problem["id"],
                stream=True,
                extra_body={"chat_template_kwargs": {"enable_thinking": True, "clear_thinking": False}},
            )

            # Collect streamed content, filtering out reasoning tokens
            parts = []
            for chunk in completion:
                if not getattr(chunk, "choices", None) or len(chunk.choices) == 0:
                    continue
                delta = getattr(chunk.choices[0], "delta", None)
                if delta is None:
                    continue
                content = getattr(delta, "content", None)
                if content is not None:
                    parts.append(content)

            raw = "".join(parts).strip()

            # Strip any accidental markdown fences
            raw = re.sub(r"^```(?:mdx|markdown)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)

            # Strip any thinking/reasoning tags
            raw = re.sub(r"<think>[\s\S]*?</think>\s*", "", raw)

            # Validate basic MDX structure
            if "---" in raw and "CodeTabs" in raw:
                return sanitize_mdx(raw)
            else:
                print(f"  Warning: Output missing MDX structure (attempt {attempt}), retrying...")
                continue

        except Exception as e:
            if "429" in str(e) and attempt < 3:
                import time
                wait = 30 * attempt
                print(f"  Rate limited (attempt {attempt}), waiting {wait}s...")
                time.sleep(wait)
                continue
            raise RuntimeError(f"NVIDIA NIM API failed after {attempt} attempt(s): {e}")

    raise RuntimeError("NVIDIA NIM failed to produce valid MDX after 3 attempts.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate a LeetCode blog post and update the Excel tracker."
    )
    parser.add_argument("problem_id", type=int, help="LeetCode problem ID (e.g. 29)")
    parser.add_argument(
        "--provider", "-p",
        default=DEFAULT_PROVIDER,
        choices=PROVIDERS,
        help=f"API provider (default: {DEFAULT_PROVIDER})",
    )
    parser.add_argument(
        "--model", "-m",
        default=DEFAULT_GEMINI_MODEL,
        choices=list(GEMINI_MODELS.keys()),
        help=f"Gemini model (only used with --provider gemini). Default: {DEFAULT_GEMINI_MODEL}",
    )
    parser.add_argument(
        "--test", action="store_true",
        help="Test mode: save to scripts/model_comparison/ instead of blog dir, skip Excel update.",
    )
    args = parser.parse_args()

    # Validate API key for chosen provider
    if args.provider == "gemini" and not os.environ.get("GEMINI_API_KEY"):
        raise SystemExit(
            "ERROR: GEMINI_API_KEY is not set.\n"
            "  1. Go to https://aistudio.google.com/apikey\n"
            "  2. Click 'Create API key'\n"
            "  3. Run: $env:GEMINI_API_KEY = 'AIza...'"
        )
    if args.provider == "nvidia" and not os.environ.get("NVIDIA_API_KEY"):
        raise SystemExit(
            "ERROR: NVIDIA_API_KEY is not set.\n"
            "  1. Go to https://build.nvidia.com\n"
            "  2. Get your API key\n"
            "  3. Run: $env:NVIDIA_API_KEY = 'nvapi-...'"
        )

    provider_label = (
        f"nvidia/{NVIDIA_MODEL}" if args.provider == "nvidia"
        else GEMINI_MODELS[args.model]
    )

    print(f"\n[1/4] Loading problem {args.problem_id} from Excel ...")
    problem = load_problem(args.problem_id)
    print(f"  Found: {problem['name']} ({problem['difficulty']})")
    print(f"  Topics: {', '.join(problem['topics']) or 'N/A'}")
    print(f"  Provider: {provider_label}")

    print("\n[2/4] Generating solutions ...")

    if args.provider == "nvidia":
        # NVIDIA generates MDX directly
        mdx_content = call_nvidia(problem)
        # Count approaches from output
        approach_count = mdx_content.count("<CodeTabs")
        print(f"  Got {approach_count} approach(es)")
    else:
        # Gemini generates JSON, then we render MDX
        doc = call_gemini(problem, model_key=args.model)
        approaches = [a["name"] for a in doc.get("Approaches", [])]
        print(f"  Got {len(approaches)} approach(es): {', '.join(approaches)}")
        mdx_content = render_mdx(doc, problem)

    print("\n[3/4] Writing blog post ...")
    slug = slugify(problem["name"])

    if args.test:
        out_dir = TEST_DIR / f"{args.problem_id}-{slug}"
        out_dir.mkdir(parents=True, exist_ok=True)
        label = args.provider if args.provider == "nvidia" else args.model
        out_file = out_dir / f"{label}.mdx"
        out_file.write_text(mdx_content, encoding="utf-8")
        print(f"  Created: {out_file.relative_to(ROOT_DIR)}")
        print(f"\n[4/4] Skipping Excel update (test mode)")
    else:
        folder_name = f"{args.problem_id}-{slug}"
        out_dir = BLOG_DIR / folder_name
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / "index.mdx"
        out_file.write_text(mdx_content, encoding="utf-8")
        print(f"  Created: {out_file.relative_to(ROOT_DIR)}")
        print("\n[4/4] Updating Excel status ...")
        mark_done(problem)

    print(f"\nDone! ({provider_label})\n")


if __name__ == "__main__":
    main()

