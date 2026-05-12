import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
import re

ROOT_DIR = Path("c:/Users/aditya/Desktop/projects/N3mo-Blog_Portfolio")
XLSX_PATH = ROOT_DIR / "scripts" / "LeetCode Questions - Last Updated 2026-05-08 05_09 CEST.xlsx"
BLOG_DIR = ROOT_DIR / "src" / "content" / "blog"

def get_done_ids():
    done_ids = set()
    for d in BLOG_DIR.iterdir():
        if d.is_dir():
            match = re.match(r"^(\d+)-", d.name)
            if match:
                done_ids.add(int(match.group(1)))
    return done_ids

def main():
    done_ids = get_done_ids()
    print(f"Found {len(done_ids)} done problems in blog directory.")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Find header row
    header_row = None
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == "ID":
            header_row = i
            break
    if header_row is None:
        print("Header row not found.")
        return
        
    status_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == "Status":
            status_col = c
            break
            
    if status_col is None:
        print("Status column not found.")
        return

    updated_count = 0
    already_done_count = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=1).value
        if raw_id is None:
            continue
        try:
            pid = int(raw_id)
        except (ValueError, TypeError):
            continue

        if pid in done_ids:
            cell = ws.cell(row=row_idx, column=status_col)
            if cell.value != "Done":
                cell.value = "Done"
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="276221", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                updated_count += 1
            else:
                already_done_count += 1
                
    wb.save(XLSX_PATH)
    print(f"Updated {updated_count} problems to Done (and {already_done_count} were already marked Done).")

if __name__ == '__main__':
    main()
