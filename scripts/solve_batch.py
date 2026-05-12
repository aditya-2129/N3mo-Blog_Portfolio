#!/usr/bin/env python3
"""
Batch-generate LeetCode blog posts in parallel using NVIDIA NIM.

Usage:
    python scripts/solve_batch.py --difficulty Easy --workers 10

Requirements:
    pip install google-generativeai openpyxl openai python-dotenv
"""

import sys
import os
import re
import time
import threading
import argparse
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── Setup paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
ROOT_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))

from dotenv import load_dotenv
load_dotenv(ROOT_DIR / ".env")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Import from solve.py
from solve import (
    XLSX_PATH, BLOG_DIR, NVIDIA_BASE_URL, NVIDIA_MODEL, NVIDIA_MAX_TOKENS,
    LANGUAGES, LANG_ATTR, slugify, extract_hyperlink_text,
    build_nvidia_prompt, call_nvidia,
)

# ── Thread-safe Excel lock ───────────────────────────────────────────────────
EXCEL_LOCK = threading.Lock()
PRINT_LOCK = threading.Lock()

# ── Stats ────────────────────────────────────────────────────────────────────
stats = {"done": 0, "failed": 0, "skipped": 0}
STATS_LOCK = threading.Lock()


def safe_print(*args, **kwargs):
    with PRINT_LOCK:
        print(*args, **kwargs)


def load_all_problems(difficulty: str = None, free_only: bool = True) -> list:
    """Load all pending problems from Excel, optionally filtered by difficulty."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Find header row
    header_row = None
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == "ID":
            header_row = i
            break
    if header_row is None:
        raise SystemExit("ERROR: Could not find header row in Excel sheet.")

    # Find Status column
    status_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == "Status":
            status_col = c
            break

    problems = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=1).value
        if raw_id is None:
            continue
        try:
            pid = int(raw_id)
        except (ValueError, TypeError):
            continue

        # Skip already done
        status = str(ws.cell(row=row_idx, column=status_col).value or "").strip()
        if status == "Done":
            continue

        diff = str(ws.cell(row=row_idx, column=7).value or "").strip()
        if difficulty and diff.lower() != difficulty.lower():
            continue

        is_free = ws.cell(row=row_idx, column=11).value == "Yes"
        if free_only and not is_free:
            continue

        name = extract_hyperlink_text(ws.cell(row=row_idx, column=2).value)
        topics_raw = ws.cell(row=row_idx, column=6).value or ""
        topics = [t.strip() for t in str(topics_raw).split(",") if t.strip()]
        accept_rate = ws.cell(row=row_idx, column=10).value

        if accept_rate:
            try:
                accept_pct = f"{float(accept_rate) * 100:.2f}%"
            except (ValueError, TypeError):
                accept_pct = str(accept_rate)
        else:
            accept_pct = "N/A"

        problems.append({
            "id": pid,
            "name": name,
            "difficulty": diff,
            "topics": topics,
            "accept_pct": accept_pct,
            "acceptance": accept_pct,
            "is_free": is_free,
            "row_idx": row_idx,
            "status_col": status_col,
        })

    wb.close()
    return problems


def mark_done_safe(problem: dict):
    """Thread-safe Excel update."""
    with EXCEL_LOCK:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        cell = ws.cell(row=problem["row_idx"], column=problem["status_col"])
        cell.value = "Done"
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(color="276221", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(XLSX_PATH)


def process_problem(problem: dict, total: int) -> tuple:
    """Generate blog post for a single problem. Returns (id, success, message)."""
    pid = problem["id"]
    name = problem["name"]

    try:
        # Generate MDX via NVIDIA NIM
        mdx_content = call_nvidia(problem)
        approach_count = mdx_content.count("<CodeTabs")

        # Write file
        slug = slugify(name)
        folder_name = f"{pid}-{slug}"
        out_dir = BLOG_DIR / folder_name
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / "index.mdx"
        out_file.write_text(mdx_content, encoding="utf-8")

        # Update Excel
        mark_done_safe(problem)

        with STATS_LOCK:
            stats["done"] += 1
            done = stats["done"]
            failed = stats["failed"]

        safe_print(f"  ✅ [{done}/{total}] #{pid} {name} — {approach_count} approaches (failed: {failed})")
        return (pid, True, f"{approach_count} approaches")

    except Exception as e:
        with STATS_LOCK:
            stats["failed"] += 1
            failed = stats["failed"]
            done = stats["done"]

        safe_print(f"  ❌ [{done}/{total}] #{pid} {name} — {e}")
        return (pid, False, str(e))


def main():
    parser = argparse.ArgumentParser(description="Batch-generate LeetCode blog posts in parallel.")
    parser.add_argument("--difficulty", "-d", default="Easy", help="Filter by difficulty (Easy/Medium/Hard)")
    parser.add_argument("--workers", "-w", type=int, default=10, help="Number of parallel workers (default: 10)")
    parser.add_argument("--limit", "-l", type=int, default=0, help="Max problems to process (0 = all)")
    parser.add_argument("--include-paid", action="store_true", help="Include paid/premium problems")
    args = parser.parse_args()

    if not os.environ.get("NVIDIA_API_KEY"):
        raise SystemExit(
            "ERROR: NVIDIA_API_KEY is not set.\n"
            "  Run: $env:NVIDIA_API_KEY = 'nvapi-...'"
        )

    print(f"\n{'='*60}")
    print(f"  LeetCode Batch Generator")
    print(f"  Model: {NVIDIA_MODEL}")
    print(f"  Difficulty: {args.difficulty}")
    print(f"  Workers: {args.workers}")
    print(f"{'='*60}\n")

    print("Loading problems from Excel...")
    problems = load_all_problems(
        difficulty=args.difficulty,
        free_only=not args.include_paid,
    )

    if args.limit > 0:
        problems = problems[:args.limit]

    total = len(problems)
    if total == 0:
        print("No pending problems found. All done!")
        return

    est_minutes = (total * 3.5) / args.workers
    est_hours = est_minutes / 60
    print(f"Found {total} pending {args.difficulty} problems")
    print(f"Estimated time: ~{est_hours:.1f} hours with {args.workers} workers")
    print(f"\nStarting at {datetime.now().strftime('%H:%M:%S')}...\n")

    start_time = time.time()
    failed_ids = []

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {}
        for i, p in enumerate(problems):
            futures[executor.submit(process_problem, p, total)] = p
            # Stagger submissions to avoid overwhelming the API
            if i < args.workers:
                time.sleep(3)  # 3s gap for initial wave

        for future in as_completed(futures):
            pid, success, msg = future.result()
            if not success:
                failed_ids.append(pid)

    elapsed = time.time() - start_time
    elapsed_min = elapsed / 60

    print(f"\n{'='*60}")
    print(f"  BATCH COMPLETE")
    print(f"  Done:    {stats['done']}")
    print(f"  Failed:  {stats['failed']}")
    print(f"  Time:    {elapsed_min:.1f} minutes")
    print(f"  Speed:   {stats['done'] / max(elapsed_min, 0.1):.1f} problems/min")
    print(f"{'='*60}")

    if failed_ids:
        print(f"\nFailed problem IDs: {failed_ids}")
        print("Re-run with: python scripts/solve_batch.py to retry")

    print()


if __name__ == "__main__":
    main()
